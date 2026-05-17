"""
Generate KoboToolbox-compatible XLSForm files for CEFANET CDF-MS.

XLSForm is the canonical spec used by KoboToolbox / ODK / SurveyCTO. It is a
plain .xlsx with three sheets: `survey`, `choices`, `settings`.

Run:  python scripts/generate_kobo_forms.py
Output: kobo-forms/*.xlsx (5 files)

Then upload each .xlsx to KoboToolbox via:
  https://kf.kobotoolbox.org/  ->  Projects -> "+ NEW" -> "Upload an XLSForm"

The forms include English labels and a `label::Nyanja (ny)` column. Edit the
Nyanja column in Kobo's UI or here before upload to refine translations.
"""
from __future__ import annotations

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

OUT_DIR = "kobo-forms"
os.makedirs(OUT_DIR, exist_ok=True)


def write_form(filename: str, settings: dict, survey: list[dict], choices: list[dict]):
    wb = Workbook()
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DCFCE7")  # ministry-100
    wrap = Alignment(wrap_text=True, vertical="top")

    # ---- survey sheet ----
    sv = wb.active
    sv.title = "survey"
    survey_cols = [
        "type", "name", "label", "label::Nyanja (ny)",
        "required", "constraint", "constraint_message", "hint",
        "appearance", "relevant", "default",
    ]
    for col_idx, header in enumerate(survey_cols, 1):
        cell = sv.cell(row=1, column=col_idx, value=header)
        cell.font = bold
        cell.fill = header_fill
    for row_idx, q in enumerate(survey, start=2):
        for col_idx, header in enumerate(survey_cols, 1):
            v = q.get(header)
            if v is None:
                continue
            cell = sv.cell(row=row_idx, column=col_idx, value=v)
            cell.alignment = wrap

    # Column widths
    widths = {
        "type": 22, "name": 22, "label": 50, "label::Nyanja (ny)": 50,
        "required": 10, "constraint": 18, "constraint_message": 30,
        "hint": 30, "appearance": 14, "relevant": 24, "default": 16,
    }
    for col_idx, header in enumerate(survey_cols, 1):
        sv.column_dimensions[chr(64 + col_idx)].width = widths.get(header, 18)

    # ---- choices sheet ----
    ch = wb.create_sheet("choices")
    choice_cols = ["list_name", "name", "label", "label::Nyanja (ny)"]
    for col_idx, header in enumerate(choice_cols, 1):
        cell = ch.cell(row=1, column=col_idx, value=header)
        cell.font = bold
        cell.fill = header_fill
    for row_idx, c in enumerate(choices, start=2):
        for col_idx, header in enumerate(choice_cols, 1):
            v = c.get(header)
            if v is None:
                continue
            ch.cell(row=row_idx, column=col_idx, value=v)
    for col_idx in range(1, len(choice_cols) + 1):
        ch.column_dimensions[chr(64 + col_idx)].width = 26

    # ---- settings sheet ----
    st = wb.create_sheet("settings")
    setting_cols = ["form_title", "form_id", "version", "default_language", "instance_name"]
    for col_idx, header in enumerate(setting_cols, 1):
        cell = st.cell(row=1, column=col_idx, value=header)
        cell.font = bold
        cell.fill = header_fill
    for col_idx, header in enumerate(setting_cols, 1):
        st.cell(row=2, column=col_idx, value=settings.get(header))

    out = os.path.join(OUT_DIR, filename)
    wb.save(out)
    print(f"  wrote {out}")


# --------------------------------------------------------------------- shared
PROJECT_STATUS_CHOICES = [
    {"list_name": "project_status", "name": "planned", "label": "Planned", "label::Nyanja (ny)": "Zokonza"},
    {"list_name": "project_status", "name": "ongoing", "label": "Ongoing", "label::Nyanja (ny)": "Zikuchitidwa"},
    {"list_name": "project_status", "name": "complete", "label": "Complete", "label::Nyanja (ny)": "Zatha"},
    {"list_name": "project_status", "name": "stalled", "label": "Stalled", "label::Nyanja (ny)": "Zayima"},
]
YES_NO = [
    {"list_name": "yes_no", "name": "yes", "label": "Yes", "label::Nyanja (ny)": "Inde"},
    {"list_name": "yes_no", "name": "no", "label": "No", "label::Nyanja (ny)": "Ayi"},
]


# ====================================================================== F-001
def project_field_update():
    settings = {
        "form_title": "Project Field Update",
        "form_id": "cefanet_project_field_update",
        "version": "2026.05.1",
        "default_language": "English (en)",
        "instance_name": "concat(${project_id}, ' · ', ${visit_date})",
    }
    survey = [
        {"type": "start", "name": "start_time"},
        {"type": "end", "name": "end_time"},
        {"type": "today", "name": "visit_date"},
        {"type": "deviceid", "name": "device_id"},
        # Visible questions
        {"type": "text", "name": "project_id", "label": "Project ID (from CDF-MS)", "label::Nyanja (ny)": "Project ID", "required": "yes", "hint": "Format: PR-001-2"},
        {"type": "select_one project_status", "name": "status", "label": "Current status", "label::Nyanja (ny)": "Mkhalidwe pano", "required": "yes"},
        {"type": "integer", "name": "completion_pct", "label": "Completion percentage", "label::Nyanja (ny)": "Peresenti ya kutha", "required": "yes", "constraint": ". >= 0 and . <= 100", "constraint_message": "Must be 0-100"},
        {"type": "integer", "name": "milestones_done", "label": "Milestones completed (out of total)", "required": "yes"},
        {"type": "text", "name": "issues", "label": "Issues encountered (narrative)", "label::Nyanja (ny)": "Mavuto akugwirizana", "hint": "Briefly describe any blockers"},
        {"type": "image", "name": "photo_current", "label": "Site photo — current state", "label::Nyanja (ny)": "Chithunzi cha malo", "required": "yes", "appearance": "new"},
        {"type": "image", "name": "photo_closeup", "label": "Photo of works (close-up)", "appearance": "new"},
        {"type": "geopoint", "name": "gps", "label": "Site GPS location", "label::Nyanja (ny)": "Malo a GPS", "required": "yes"},
        {"type": "text", "name": "supervisor", "label": "Supervisor on site (full name)", "required": "yes"},
        {"type": "text", "name": "officer_name", "label": "Your name (field officer)", "required": "yes"},
    ]
    write_form("01_project_field_update.xlsx", settings, survey, PROJECT_STATUS_CHOICES + YES_NO)


# ====================================================================== F-002
def community_scorecard():
    settings = {
        "form_title": "Community Scorecard",
        "form_id": "cefanet_community_scorecard",
        "version": "2026.05.1",
        "default_language": "English (en)",
        "instance_name": "concat(${facility_name}, ' · ', ${session_date})",
    }
    survey = [
        {"type": "start", "name": "start_time"},
        {"type": "end", "name": "end_time"},
        {"type": "today", "name": "session_date"},
        {"type": "text", "name": "facility_name", "label": "Facility / project being scored", "label::Nyanja (ny)": "Dzina la malo amene mukuwerenga", "required": "yes"},
        {"type": "select_one facility_category", "name": "category", "label": "Category", "required": "yes"},
        {"type": "select_one province", "name": "province", "label": "Province", "required": "yes"},
        {"type": "text", "name": "constituency", "label": "Constituency", "required": "yes"},
        {"type": "text", "name": "ward", "label": "Ward", "required": "yes"},
        {"type": "integer", "name": "participants_total", "label": "Number of community participants", "label::Nyanja (ny)": "Chiwerengero cha anthu", "required": "yes", "constraint": ". >= 10", "constraint_message": "Minimum 10 participants for a valid scorecard"},
        {"type": "integer", "name": "participants_women", "label": "Number of female participants", "required": "yes"},
        {"type": "integer", "name": "participants_youth", "label": "Number of youth (≤ 35 yrs) participants", "required": "yes"},
        {"type": "integer", "name": "score_access", "label": "Access score (0-100)", "label::Nyanja (ny)": "Mwa nu kuyandikira", "required": "yes", "constraint": ". >= 0 and . <= 100"},
        {"type": "integer", "name": "score_quality", "label": "Quality score (0-100)", "required": "yes", "constraint": ". >= 0 and . <= 100"},
        {"type": "integer", "name": "score_provider", "label": "Provider behaviour score (0-100)", "required": "yes", "constraint": ". >= 0 and . <= 100"},
        {"type": "integer", "name": "score_satisfaction", "label": "Satisfaction score (0-100)", "required": "yes", "constraint": ". >= 0 and . <= 100"},
        {"type": "text", "name": "narrative", "label": "Community narrative — what was said?", "label::Nyanja (ny)": "Nkhani ya anthu", "hint": "Quotes and observations from the session"},
        {"type": "select_multiple priority_actions", "name": "priorities", "label": "Top priority actions identified", "appearance": "minimal"},
        {"type": "image", "name": "photo_session", "label": "Photo of the scorecard session", "appearance": "new"},
        {"type": "geopoint", "name": "gps", "label": "Session GPS"},
        {"type": "text", "name": "submitted_by", "label": "Submitted by (WDC member name)", "required": "yes"},
    ]
    choices = [
        {"list_name": "facility_category", "name": "health", "label": "Health facility"},
        {"list_name": "facility_category", "name": "education", "label": "Education facility"},
        {"list_name": "facility_category", "name": "infrastructure", "label": "Infrastructure"},
        {"list_name": "facility_category", "name": "empowerment", "label": "Empowerment / grants"},
        {"list_name": "priority_actions", "name": "drugs", "label": "Drug stock-out resolution"},
        {"list_name": "priority_actions", "name": "staff", "label": "Increase staff on post"},
        {"list_name": "priority_actions", "name": "infrastructure", "label": "Infrastructure repair"},
        {"list_name": "priority_actions", "name": "outreach", "label": "Community outreach"},
        {"list_name": "priority_actions", "name": "training", "label": "Training / capacity building"},
    ] + [
        {"list_name": "province", "name": p.lower().replace(" ", "_").replace("-", "_"), "label": p}
        for p in ["Central", "Copperbelt", "Eastern", "Luapula", "Lusaka", "Muchinga", "Northern", "North-Western", "Southern", "Western"]
    ] + YES_NO
    write_form("02_community_scorecard.xlsx", settings, survey, choices)


# ====================================================================== F-003
def grievance_intake():
    settings = {
        "form_title": "Grievance Intake",
        "form_id": "cefanet_grievance_intake",
        "version": "2026.05.1",
        "default_language": "English (en)",
        "instance_name": "concat('Grievance · ', ${logged_at})",
    }
    survey = [
        {"type": "start", "name": "start_time"},
        {"type": "end", "name": "end_time"},
        {"type": "today", "name": "logged_at"},
        {"type": "select_one yes_no", "name": "anonymous", "label": "Submit this grievance anonymously?", "label::Nyanja (ny)": "Mukufuna kuti mukhale mosadziwika?", "required": "yes"},
        {"type": "select_one grievance_category", "name": "category", "label": "What is the grievance about?", "required": "yes"},
        {"type": "select_one province", "name": "province", "label": "Province", "required": "yes"},
        {"type": "text", "name": "constituency", "label": "Constituency", "required": "yes"},
        {"type": "text", "name": "description", "label": "Describe the issue", "label::Nyanja (ny)": "Fotokozani vuto", "required": "yes", "hint": "Be as specific as you can"},
        {"type": "image", "name": "photo_evidence", "label": "Photo evidence (optional)", "appearance": "new"},
        {"type": "text", "name": "complainant_name", "label": "Your name (kept confidential)", "relevant": "${anonymous} = 'no'"},
        {"type": "text", "name": "complainant_phone", "label": "Phone number (kept confidential)", "relevant": "${anonymous} = 'no'", "constraint": "regex(., '^[0-9]{10,12}$')", "constraint_message": "Enter a valid phone number"},
        {"type": "text", "name": "witness", "label": "Witness name (optional)"},
    ]
    choices = [
        {"list_name": "grievance_category", "name": "service_quality", "label": "Service quality"},
        {"list_name": "grievance_category", "name": "fund_misuse", "label": "Suspected fund misuse"},
        {"list_name": "grievance_category", "name": "missing_beneficiary", "label": "Eligible beneficiary excluded"},
        {"list_name": "grievance_category", "name": "delay", "label": "Project delay"},
        {"list_name": "grievance_category", "name": "procurement", "label": "Procurement concern"},
    ] + [
        {"list_name": "province", "name": p.lower().replace(" ", "_").replace("-", "_"), "label": p}
        for p in ["Central", "Copperbelt", "Eastern", "Luapula", "Lusaka", "Muchinga", "Northern", "North-Western", "Southern", "Western"]
    ] + YES_NO
    write_form("03_grievance_intake.xlsx", settings, survey, choices)


# ====================================================================== F-004
def bursary_verification():
    settings = {
        "form_title": "Bursary Beneficiary Verification",
        "form_id": "cefanet_bursary_verification",
        "version": "2026.05.1",
        "default_language": "English (en)",
        "instance_name": "concat(${beneficiary_code}, ' · ', ${term})",
    }
    survey = [
        {"type": "start", "name": "start_time"},
        {"type": "end", "name": "end_time"},
        {"type": "today", "name": "visit_date"},
        {"type": "text", "name": "beneficiary_code", "label": "Beneficiary code (BEN-####)", "required": "yes", "constraint": "regex(., '^BEN-[0-9]{4}$')"},
        {"type": "integer", "name": "nrc_last4", "label": "Last 4 digits of NRC", "required": "yes", "constraint": ". >= 0 and . <= 9999"},
        {"type": "select_one yes_no", "name": "enrolled", "label": "Currently enrolled in school?", "required": "yes"},
        {"type": "select_one term", "name": "term", "label": "Verification for which term?", "required": "yes"},
        {"type": "integer", "name": "attendance_pct", "label": "Attendance percentage (Term)", "constraint": ". >= 0 and . <= 100", "relevant": "${enrolled} = 'yes'"},
        {"type": "decimal", "name": "grade_avg", "label": "Grade average (out of 100)", "constraint": ". >= 0 and . <= 100", "relevant": "${enrolled} = 'yes'"},
        {"type": "image", "name": "photo_report", "label": "Photo of report card / acceptance letter", "appearance": "new", "relevant": "${enrolled} = 'yes'"},
        {"type": "text", "name": "school", "label": "School / institution name", "required": "yes"},
        {"type": "text", "name": "verifier", "label": "Verifier signature (full name)", "required": "yes"},
    ]
    choices = [
        {"list_name": "term", "name": "term_1", "label": "Term 1"},
        {"list_name": "term", "name": "term_2", "label": "Term 2"},
        {"list_name": "term", "name": "term_3", "label": "Term 3"},
    ] + YES_NO
    write_form("04_bursary_verification.xlsx", settings, survey, choices)


# ====================================================================== F-005
def health_facility_inspection():
    settings = {
        "form_title": "Health Facility Inspection",
        "form_id": "cefanet_health_inspection",
        "version": "2026.05.1",
        "default_language": "English (en)",
        "instance_name": "concat(${facility_name}, ' · ', ${visit_date})",
    }
    survey = [
        {"type": "start", "name": "start_time"},
        {"type": "end", "name": "end_time"},
        {"type": "today", "name": "visit_date"},
        {"type": "text", "name": "facility_name", "label": "Facility name", "required": "yes"},
        {"type": "select_one facility_type", "name": "facility_type", "label": "Facility type", "required": "yes"},
        {"type": "select_one operational_status", "name": "status", "label": "Operational status today", "required": "yes"},
        {"type": "integer", "name": "drug_availability", "label": "Drug availability (0-100%)", "required": "yes", "constraint": ". >= 0 and . <= 100"},
        {"type": "select_one yes_no", "name": "water", "label": "Has running water right now?", "required": "yes"},
        {"type": "select_one yes_no", "name": "electricity", "label": "Has electricity right now?", "required": "yes"},
        {"type": "integer", "name": "staff_on_post", "label": "Staff on post today", "required": "yes"},
        {"type": "integer", "name": "staff_established", "label": "Staff establishment (positions)", "required": "yes"},
        {"type": "integer", "name": "patients_today", "label": "Patients seen so far today", "required": "yes"},
        {"type": "integer", "name": "anc_visits_month", "label": "ANC visits this month"},
        {"type": "integer", "name": "immunisations_month", "label": "Children immunised this month"},
        {"type": "image", "name": "photo_pharmacy", "label": "Photo — pharmacy shelves", "required": "yes", "appearance": "new"},
        {"type": "image", "name": "photo_register", "label": "Photo — outpatient register", "appearance": "new"},
        {"type": "geopoint", "name": "gps", "label": "Facility GPS", "required": "yes"},
        {"type": "text", "name": "inspector", "label": "Inspector name", "required": "yes"},
        {"type": "text", "name": "in_charge", "label": "Facility in-charge name", "required": "yes"},
    ]
    choices = [
        {"list_name": "facility_type", "name": "clinic", "label": "Clinic"},
        {"list_name": "facility_type", "name": "health_post", "label": "Health post"},
        {"list_name": "facility_type", "name": "maternity_shelter", "label": "Maternity shelter"},
        {"list_name": "facility_type", "name": "borehole", "label": "Borehole / water point"},
        {"list_name": "facility_type", "name": "staff_housing", "label": "Staff housing"},
        {"list_name": "operational_status", "name": "operational", "label": "Fully operational"},
        {"list_name": "operational_status", "name": "partial", "label": "Partially operational"},
        {"list_name": "operational_status", "name": "non_operational", "label": "Non-operational"},
        {"list_name": "operational_status", "name": "under_construction", "label": "Under construction"},
    ] + YES_NO
    write_form("05_health_facility_inspection.xlsx", settings, survey, choices)


# README inside kobo-forms/
def write_readme():
    out = os.path.join(OUT_DIR, "README.md")
    with open(out, "w", encoding="utf-8") as f:
        f.write("""# CEFANET CDF-MS — Kobo Forms

Five XLSForm files ready for upload to KoboToolbox.

## How to upload

1. Open https://kf.kobotoolbox.org/ and sign in.
2. Top right → **+ NEW** → **Upload an XLSForm**.
3. Pick one of the five `.xlsx` files in this folder.
4. Click **Deploy** on the resulting project.
5. From the project's **Settings** → **Form** panel, copy the **Asset UID**
   (looks like `aXrFt7HsM9zG3kQv2bN8cV`).
6. Paste each UID into the repo's `.env`:

```
KOBO_BASE=https://kf.kobotoolbox.org
KOBO_TOKEN=<your-api-token>
KOBO_UID_PROJECT_UPDATE=<uid>
KOBO_UID_SCORECARD=<uid>
KOBO_UID_GRIEVANCE=<uid>
KOBO_UID_BURSARY=<uid>
KOBO_UID_HEALTH=<uid>
```

Then rebuild and redeploy. Live submissions flow into `/forms`.

## Files

| # | File | Purpose |
|---|------|---------|
| 1 | `01_project_field_update.xlsx`     | CDFC officer monthly project status check |
| 2 | `02_community_scorecard.xlsx`      | WDC-led 4-dimension community assessment |
| 3 | `03_grievance_intake.xlsx`         | Citizen grievance logging |
| 4 | `04_bursary_verification.xlsx`     | Quarterly bursary beneficiary check |
| 5 | `05_health_facility_inspection.xlsx` | Monthly CDF-funded health facility inspection |
""")


def main():
    print("Generating XLSForm files in kobo-forms/...")
    project_field_update()
    community_scorecard()
    grievance_intake()
    bursary_verification()
    health_facility_inspection()
    write_readme()
    print("Done.")


if __name__ == "__main__":
    main()
